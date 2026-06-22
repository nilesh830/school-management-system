"""
SMS-060 — Export Reports to PDF / Excel.

For each report (attendance / grades / fees):
  - PDF  export → 200, application/pdf, non-empty bytes, attachment header
  - Excel export → 200, xlsx mimetype, non-empty bytes, header row loadable
  - invalid format → 400
  - RBAC: parent / student forbidden; teacher forbidden on the fees export

xhtml2pdf is mocked (same stub the report-card / fee-receipt tests use) so the
PDF path exercises template rendering + data assembly without real rendering.
Excel is generated for real — openpyxl is pure Python.
"""
from io import BytesIO
from datetime import date, timedelta
from unittest.mock import patch, MagicMock

import openpyxl

from app.models.user import User
from app.models.student import Student
from app.models.class_ import Class
from app.models.section import Section
from app.models.academic_year import AcademicYear
from app.models.attendance import Attendance
from app.models.exam import Exam
from app.models.exam_result import ExamResult
from app.models.subject import Subject
from app.models.fee_structure import FeeStructure
from app.models.fee_record import FeeRecord
from app.models.fee_payment import FeePayment


_XLSX_MIMETYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)

_counter = 0


def _uid():
    global _counter
    _counter += 1
    return _counter


def _fake_pisa(html, dest, **kwargs):
    """xhtml2pdf stub: writes a fake PDF signature and reports no errors."""
    dest.write(b'%PDF-FAKE')
    result = MagicMock()
    result.err = 0
    return result


# ---------------------------------------------------------------------------
# Shared factories (mirrors test_reports.py)
# ---------------------------------------------------------------------------

def make_class(db, grade_level=1):
    c = Class(name=f'Grade Exp {_uid()}', grade_level=grade_level)
    db.session.add(c)
    db.session.commit()
    return c


def make_section(db, class_id):
    uid = _uid()
    sec = Section(name=f'Exp{uid}', class_id=class_id, is_active=True)
    db.session.add(sec)
    db.session.commit()
    return sec


def make_academic_year(db):
    uid = _uid()
    ay = AcademicYear(name=f'AY-Exp-{uid}', start_date=date(2024, 6, 1),
                      end_date=date(2025, 5, 31), is_current=True, is_active=True)
    db.session.add(ay)
    db.session.commit()
    return ay


def make_student(db):
    uid = _uid()
    u = User(email=f'expstu_{uid}@test.sms', role='student',
             first_name=f'Exp{uid}', last_name='Stu')
    u.set_password('Student@123')
    db.session.add(u)
    db.session.flush()
    s = Student(user_id=u.id, admission_no=f'ADM-EXP-{uid:05d}',
                first_name=f'Exp{uid}', last_name='Stu',
                date_of_birth=date(2011, 1, 1), gender='Male',
                admission_date=date(2024, 6, 1))
    db.session.add(s)
    db.session.commit()
    return s


def make_subject(db, max_marks=100):
    uid = _uid()
    sub = Subject(name=f'Subject Exp {uid}', code=f'EXP{uid:04d}', max_marks=max_marks)
    db.session.add(sub)
    db.session.commit()
    return sub


def make_exam(db, section_id, academic_year_id):
    uid = _uid()
    e = Exam(name=f'Exam Exp {uid}', term='Term 1', exam_type='midterm',
             section_id=section_id, academic_year_id=academic_year_id, is_active=True)
    db.session.add(e)
    db.session.commit()
    return e


def add_result(db, exam_id, student_id, subject_id, marks, grade, gpa):
    db.session.add(ExamResult(
        exam_id=exam_id, student_id=student_id, subject_id=subject_id,
        marks_obtained=marks, grade=grade, gpa=gpa, status='finalized',
    ))
    db.session.commit()


def make_student_token(client, db):
    """Create a fresh student user and return its access token."""
    uid = _uid()
    email = f'exptok_{uid}@test.sms'
    u = User(email=email, role='student', first_name='ExpTok', last_name='Stu')
    u.set_password('Student@123')
    db.session.add(u)
    db.session.commit()
    r = client.post('/api/v1/auth/login',
                    json={'email': email, 'password': 'Student@123',
                          'school_slug': 'test'})
    return r.get_json()['data']['access_token']


# ---------------------------------------------------------------------------
# Scenario seeders
# ---------------------------------------------------------------------------

def seed_attendance(db):
    cls = make_class(db)
    sec = make_section(db, cls.id)
    s1 = make_student(db)
    base = date(2025, 1, 6)
    for i, st in enumerate(['present', 'present', 'absent']):
        db.session.add(Attendance(student_id=s1.id, section_id=sec.id,
                                  date=base + timedelta(days=i), status=st))
    db.session.commit()
    return sec


def seed_grades(db):
    cls = make_class(db)
    sec = make_section(db, cls.id)
    ay = make_academic_year(db)
    exam = make_exam(db, sec.id, ay.id)
    sub = make_subject(db, max_marks=100)
    s1 = make_student(db)
    add_result(db, exam.id, s1.id, sub.id, 95, 'A+', 4.0)
    return exam


def seed_fees(db):
    cls = make_class(db)
    ay = make_academic_year(db)
    student = make_student(db)
    fs = FeeStructure(class_id=cls.id, academic_year_id=ay.id,
                      fee_type='TuitionExp', amount=5000.00,
                      is_recurring=False, frequency='one_time')
    db.session.add(fs)
    db.session.commit()
    fr = FeeRecord(student_id=student.id, fee_structure_id=fs.id,
                   amount=5000.00, discount=0, net_amount=5000.00,
                   due_date=date.today() - timedelta(days=2), status='partial')
    db.session.add(fr)
    db.session.commit()
    db.session.add(FeePayment(fee_record_id=fr.id, amount_paid=2000.00,
                              payment_method='cash', payment_date=date.today(),
                              receipt_no=f'REC-EXP-{_uid():06d}'))
    db.session.commit()
    return cls


def _attendance_url(sec, fmt):
    return (f'/api/v1/reports/attendance/export?format={fmt}'
            f'&section_id={sec.id}&from_date=2025-01-01&to_date=2025-01-31')


# ===========================================================================
# Attendance export
# ===========================================================================

class TestAttendanceExport:

    def test_pdf(self, client, admin_token, db):
        sec = seed_attendance(db)
        with patch('xhtml2pdf.pisa.CreatePDF', side_effect=_fake_pisa):
            resp = client.get(_attendance_url(sec, 'pdf'),
                              headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 200
        assert 'application/pdf' in resp.content_type
        assert len(resp.data) > 0
        assert 'attachment' in resp.headers.get('Content-Disposition', '')

    def test_excel(self, client, admin_token, db):
        sec = seed_attendance(db)
        resp = client.get(_attendance_url(sec, 'excel'),
                          headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 200
        assert resp.content_type.startswith(_XLSX_MIMETYPE)
        assert len(resp.data) > 0
        assert 'attachment' in resp.headers.get('Content-Disposition', '')
        wb = openpyxl.load_workbook(BytesIO(resp.data))
        ws = wb.active
        header = [c.value for c in ws[1]]
        assert header[0] == 'Student'
        assert 'Percentage' in header

    def test_default_format_is_pdf(self, client, admin_token, db):
        sec = seed_attendance(db)
        url = (f'/api/v1/reports/attendance/export?section_id={sec.id}'
               f'&from_date=2025-01-01&to_date=2025-01-31')
        with patch('xhtml2pdf.pisa.CreatePDF', side_effect=_fake_pisa):
            resp = client.get(url, headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 200
        assert 'application/pdf' in resp.content_type

    def test_invalid_format_400(self, client, admin_token, db):
        sec = seed_attendance(db)
        resp = client.get(_attendance_url(sec, 'csv'),
                          headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 400

    def test_teacher_allowed(self, client, teacher_token, db):
        sec = seed_attendance(db)
        with patch('xhtml2pdf.pisa.CreatePDF', side_effect=_fake_pisa):
            resp = client.get(_attendance_url(sec, 'pdf'),
                              headers={'Authorization': f'Bearer {teacher_token}'})
        assert resp.status_code == 200

    def test_parent_forbidden(self, client, parent_token, db):
        sec = seed_attendance(db)
        resp = client.get(_attendance_url(sec, 'pdf'),
                          headers={'Authorization': f'Bearer {parent_token}'})
        assert resp.status_code == 403

    def test_student_forbidden(self, client, db):
        sec = seed_attendance(db)
        token = make_student_token(client, db)
        resp = client.get(_attendance_url(sec, 'pdf'),
                          headers={'Authorization': f'Bearer {token}'})
        assert resp.status_code == 403

    def test_missing_section_404(self, client, admin_token, db):
        url = ('/api/v1/reports/attendance/export?format=pdf&section_id=999999'
               '&from_date=2025-01-01&to_date=2025-01-31')
        with patch('xhtml2pdf.pisa.CreatePDF', side_effect=_fake_pisa):
            resp = client.get(url, headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 404


# ===========================================================================
# Grades export
# ===========================================================================

class TestGradesExport:

    def test_pdf(self, client, admin_token, db):
        exam = seed_grades(db)
        with patch('xhtml2pdf.pisa.CreatePDF', side_effect=_fake_pisa):
            resp = client.get(f'/api/v1/reports/grades/export?format=pdf&exam_id={exam.id}',
                              headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 200
        assert 'application/pdf' in resp.content_type
        assert len(resp.data) > 0
        assert 'attachment' in resp.headers.get('Content-Disposition', '')

    def test_excel(self, client, admin_token, db):
        exam = seed_grades(db)
        resp = client.get(f'/api/v1/reports/grades/export?format=excel&exam_id={exam.id}',
                          headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 200
        assert resp.content_type.startswith(_XLSX_MIMETYPE)
        assert len(resp.data) > 0
        wb = openpyxl.load_workbook(BytesIO(resp.data))
        header = [c.value for c in wb.active[1]]
        assert header[0] == 'Student'
        assert 'Grade' in header

    def test_invalid_format_400(self, client, admin_token, db):
        exam = seed_grades(db)
        resp = client.get(f'/api/v1/reports/grades/export?format=xml&exam_id={exam.id}',
                          headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 400

    def test_missing_exam_404(self, client, admin_token, db):
        with patch('xhtml2pdf.pisa.CreatePDF', side_effect=_fake_pisa):
            resp = client.get('/api/v1/reports/grades/export?format=pdf&exam_id=999999',
                              headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 404

    def test_missing_exam_id_400(self, client, admin_token, db):
        resp = client.get('/api/v1/reports/grades/export?format=pdf',
                          headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 400

    def test_teacher_allowed(self, client, teacher_token, db):
        exam = seed_grades(db)
        resp = client.get(f'/api/v1/reports/grades/export?format=excel&exam_id={exam.id}',
                          headers={'Authorization': f'Bearer {teacher_token}'})
        assert resp.status_code == 200

    def test_parent_forbidden(self, client, parent_token, db):
        resp = client.get('/api/v1/reports/grades/export?format=pdf&exam_id=1',
                          headers={'Authorization': f'Bearer {parent_token}'})
        assert resp.status_code == 403

    def test_student_forbidden(self, client, db):
        token = make_student_token(client, db)
        resp = client.get('/api/v1/reports/grades/export?format=pdf&exam_id=1',
                          headers={'Authorization': f'Bearer {token}'})
        assert resp.status_code == 403


# ===========================================================================
# Fees export
# ===========================================================================

class TestFeesExport:

    def test_pdf(self, client, admin_token, db):
        seed_fees(db)
        with patch('xhtml2pdf.pisa.CreatePDF', side_effect=_fake_pisa):
            resp = client.get('/api/v1/reports/fees/export?format=pdf',
                              headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 200
        assert 'application/pdf' in resp.content_type
        assert len(resp.data) > 0
        assert 'attachment' in resp.headers.get('Content-Disposition', '')

    def test_excel(self, client, admin_token, db):
        seed_fees(db)
        resp = client.get('/api/v1/reports/fees/export?format=excel',
                          headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 200
        assert resp.content_type.startswith(_XLSX_MIMETYPE)
        assert len(resp.data) > 0
        wb = openpyxl.load_workbook(BytesIO(resp.data))
        header = [c.value for c in wb.active[1]]
        assert header[0] == 'Fee Type'
        assert 'Collected' in header

    def test_default_format_is_pdf(self, client, admin_token, db):
        seed_fees(db)
        with patch('xhtml2pdf.pisa.CreatePDF', side_effect=_fake_pisa):
            resp = client.get('/api/v1/reports/fees/export',
                              headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 200
        assert 'application/pdf' in resp.content_type

    def test_invalid_format_400(self, client, admin_token, db):
        resp = client.get('/api/v1/reports/fees/export?format=json',
                          headers={'Authorization': f'Bearer {admin_token}'})
        assert resp.status_code == 400

    def test_teacher_forbidden(self, client, teacher_token, db):
        resp = client.get('/api/v1/reports/fees/export?format=pdf',
                          headers={'Authorization': f'Bearer {teacher_token}'})
        assert resp.status_code == 403

    def test_parent_forbidden(self, client, parent_token, db):
        resp = client.get('/api/v1/reports/fees/export?format=pdf',
                          headers={'Authorization': f'Bearer {parent_token}'})
        assert resp.status_code == 403

    def test_student_forbidden(self, client, db):
        token = make_student_token(client, db)
        resp = client.get('/api/v1/reports/fees/export?format=pdf',
                          headers={'Authorization': f'Bearer {token}'})
        assert resp.status_code == 403
