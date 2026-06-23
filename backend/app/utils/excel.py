"""
Generic Excel (.xlsx) generation helper for report exports (SMS-060).

Pure-Python via openpyxl — no system libraries required. The single public
function build_xlsx() takes a sheet title, a list of column headers, and a list
of row value-lists, and returns the workbook as raw xlsx bytes (in a BytesIO).

Reusable across all report exporters; contains no business logic.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# Header styling — dark fill + white bold text to match the PDF table headers.
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


def build_xlsx(sheet_title: str, headers: list, rows: list) -> bytes:
    """
    Build an .xlsx workbook with a single sheet and return its bytes.

    Args:
        sheet_title: Worksheet name (truncated to 31 chars — Excel's limit).
        headers:     List of column header strings (the first row).
        rows:        List of row value-lists; each inner list is one row.

    Returns:
        The workbook serialized to xlsx as bytes.
    """
    wb = Workbook()
    ws = wb.active
    # Excel caps sheet titles at 31 characters.
    ws.title = (sheet_title or "Report")[:31]

    # Header row
    ws.append(list(headers))
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # Data rows
    for row in rows:
        ws.append(list(row))

    # Best-effort column auto-width based on the longest value per column.
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                value = row[col_idx - 1]
                max_len = max(max_len, len(str(value)) if value is not None else 0)
        # +2 padding, capped so a stray long value doesn't blow out the sheet.
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
